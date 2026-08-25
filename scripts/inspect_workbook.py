import os
from openpyxl import load_workbook

path = r'C:\Users\USER\Downloads\NeedyNeedsDatabase.xlsx'
print(f'EXISTS={os.path.exists(path)}')
wb = load_workbook(path, data_only=True)
print('SHEETS=' + ', '.join(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    max_row = ws.max_row
    max_col = ws.max_column
    print(f'--- SHEET {name} rows={max_row} cols={max_col}')
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 8), values_only=True):
        print(row)
